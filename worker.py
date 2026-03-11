import argparse
import json
import os
import csv
from datetime import datetime

import cv2
import numpy as np
from fpdf import FPDF
from openpyxl import Workbook

import Nut_Detection
import Bolt_Detection
import Gear_Detection

VALID_CATEGORIES = {"All", "Nut", "Bolt", "Gear"}


def _timestamp():
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


def _save_image(frame, outputs_dir, prefix):
    name = f"{prefix}_{_timestamp()}.jpg"
    path = os.path.join(outputs_dir, name)
    cv2.imwrite(path, frame)
    return path, name


def _run_detection_on_frame(category, frame, conf=None, iou=None, label_mode="confidence"):
    if category == "Nut":
        return Nut_Detection.detect_frame(frame, conf, iou, label_mode)
    if category == "Bolt":
        return Bolt_Detection.detect_frame(frame, conf, iou, label_mode)
    if category == "Gear":
        frame, detected, defect, _ = Gear_Detection.detect_frame(frame, conf, iou, label_mode)
        return frame, detected, defect
    raise ValueError("Unsupported category")


def _run_detection_all(frame, conf=None, iou=None, label_mode="confidence"):
    out = frame
    n_frame, n_det, n_def = Nut_Detection.detect_frame(out, conf, iou, label_mode, draw_counts=False)
    out = n_frame
    b_frame, b_det, b_def = Bolt_Detection.detect_frame(out, conf, iou, label_mode, draw_counts=False)
    out = b_frame
    g_frame, g_det, g_def, _ = Gear_Detection.detect_frame(out, conf, iou, label_mode, draw_counts=False)
    out = g_frame
    totals = (n_det + b_det + g_det, n_def + b_def + g_def)
    per_category = {
        "Nut": {"detected": n_det, "defect": n_def},
        "Bolt": {"detected": b_det, "defect": b_def},
        "Gear": {"detected": g_det, "defect": g_def},
    }
    return out, totals, per_category


def process_image(payload):
    image_path = payload["path"]
    category = payload["category"]
    conf = payload.get("conf")
    iou = payload.get("iou")
    label_mode = payload.get("label_mode") or "confidence"
    outputs_dir = payload["outputs_dir"]

    frame = cv2.imread(image_path)
    if frame is None:
        return {"ok": False, "error": "Invalid image"}

    if category == "All":
        processed, (detected, defect), per_category = _run_detection_all(frame, conf, iou, label_mode)
        good = max(detected - defect, 0)
        output_path, output_name = _save_image(processed, outputs_dir, "image_all")
        return {
            "ok": True,
            "detected": detected,
            "defect": defect,
            "good": good,
            "output_url": f"/outputs/{output_name}",
            "per_category": per_category,
        }

    processed, detected, defect = _run_detection_on_frame(category, frame, conf, iou, label_mode)
    good = max(detected - defect, 0)
    output_path, output_name = _save_image(processed, outputs_dir, f"image_{category.lower()}")
    return {
        "ok": True,
        "detected": detected,
        "defect": defect,
        "good": good,
        "output_url": f"/outputs/{output_name}",
    }


def process_video(payload):
    video_path = payload["path"]
    category = payload["category"]
    outputs_dir = payload["outputs_dir"]

    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        return {"ok": False, "error": "Failed to open video"}

    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    fps = cap.get(cv2.CAP_PROP_FPS) or 24

    output_name = f"output_{category.lower()}_{_timestamp()}.mp4"
    output_path = os.path.join(outputs_dir, output_name)

    fourcc = cv2.VideoWriter_fourcc(*"mp4v")
    writer = cv2.VideoWriter(output_path, fourcc, fps, (width, height))

    detected_total = 0
    defect_total = 0
    per_category_totals = {
        "Nut": {"detected": 0, "defect": 0},
        "Bolt": {"detected": 0, "defect": 0},
        "Gear": {"detected": 0, "defect": 0},
    }
    last_frame = None

    while True:
        ret, frame = cap.read()
        if not ret:
            break
        if category == "All":
            processed, (detected, defect), per_category = _run_detection_all(frame)
            for cat, vals in per_category.items():
                per_category_totals[cat]["detected"] += vals["detected"]
                per_category_totals[cat]["defect"] += vals["defect"]
        else:
            processed, detected, defect = _run_detection_on_frame(category, frame)

        detected_total += detected
        defect_total += defect
        writer.write(processed)
        last_frame = processed

    cap.release()
    writer.release()

    good = max(detected_total - defect_total, 0)
    result = {
        "ok": True,
        "detected": detected_total,
        "defect": defect_total,
        "good": good,
        "output_url": f"/outputs/{output_name}",
    }
    if category == "All":
        result["per_category"] = per_category_totals

    if last_frame is not None:
        _save_image(last_frame, outputs_dir, f"video_{category.lower()}")

    return result


def _read_log(log_path):
    rows = []
    if not os.path.exists(log_path):
        return rows
    try:
        with open(log_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(row)
    except Exception:
        return rows
    return rows


def _write_pdf_report(path, title, rows, totals=None, date_label=None):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, title, ln=True)
    pdf.set_font("Arial", size=11)
    if date_label:
        pdf.cell(0, 8, f"Date: {date_label}", ln=True)
    pdf.ln(2)

    pdf.set_font("Arial", "B", 10)
    pdf.cell(40, 8, "Timestamp", 1)
    pdf.cell(30, 8, "Category", 1)
    pdf.cell(30, 8, "Detected", 1)
    pdf.cell(25, 8, "Defect", 1)
    pdf.cell(25, 8, "Good", 1)
    pdf.ln()

    pdf.set_font("Arial", size=10)
    for row in rows:
        pdf.cell(40, 8, row.get("timestamp", ""), 1)
        pdf.cell(30, 8, row.get("category", ""), 1)
        pdf.cell(30, 8, str(row.get("detected", "")), 1)
        pdf.cell(25, 8, str(row.get("defect", "")), 1)
        pdf.cell(25, 8, str(row.get("good", "")), 1)
        pdf.ln()

    if totals:
        pdf.ln(4)
        pdf.set_font("Arial", "B", 11)
        pdf.cell(0, 8, "Totals", ln=True)
        pdf.set_font("Arial", size=10)
        for k, v in totals.items():
            pdf.cell(0, 7, f"{k}: {v}", ln=True)

    pdf.output(path)


def _write_excel_report(path, rows, date_label=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append(["Timestamp", "Category", "Detected", "Defect", "Good"])
    for row in rows:
        ws.append([
            row.get("timestamp", ""),
            row.get("category", ""),
            int(row.get("detected", 0)),
            int(row.get("defect", 0)),
            int(row.get("good", 0))
        ])
    wb.save(path)


def report_latest(payload):
    log_path = payload["log_path"]
    reports_dir = payload["reports_dir"]
    rows = _read_log(log_path)
    if not rows:
        return {"ok": False, "error": "No logs"}
    latest = rows[-1:]
    out_name = f"report_latest_{_timestamp()}.pdf"
    out_path = os.path.join(reports_dir, out_name)
    _write_pdf_report(out_path, "Latest Detection Report", latest)
    return {"ok": True, "path": out_path, "url": f"/reports/{out_name}"}


def report_daily(payload, kind):
    log_path = payload["log_path"]
    reports_dir = payload["reports_dir"]
    date = payload.get("date")
    rows = _read_log(log_path)
    if not rows:
        return {"ok": False, "error": "No logs"}
    if date:
        rows = [r for r in rows if (r.get("timestamp") or "").startswith(date)]
    if not rows:
        return {"ok": False, "error": "No logs for date"}
    totals = {
        "Detected": sum(int(r.get("detected", 0)) for r in rows),
        "Defect": sum(int(r.get("defect", 0)) for r in rows),
        "Good": sum(int(r.get("good", 0)) for r in rows),
    }
    date_label = date or datetime.now().strftime("%Y-%m-%d")
    if kind == "pdf":
        out_name = f"report_daily_{date_label}_{_timestamp()}.pdf"
        out_path = os.path.join(reports_dir, out_name)
        _write_pdf_report(out_path, "Daily Report", rows, totals=totals, date_label=date_label)
        return {"ok": True, "path": out_path, "url": f"/reports/{out_name}"}
    out_name = f"report_daily_{date_label}_{_timestamp()}.xlsx"
    out_path = os.path.join(reports_dir, out_name)
    _write_excel_report(out_path, rows, date_label=date_label)
    return {"ok": True, "path": out_path, "url": f"/reports/{out_name}"}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--payload", required=True)
    args = parser.parse_args()

    payload = json.loads(args.payload)
    mode = payload.get("mode")
    category = payload.get("category", "Nut")

    if mode in ("image", "video"):
        if category not in VALID_CATEGORIES:
            print(json.dumps({"ok": False, "error": "Invalid category"}))
            return

    if mode == "image":
        result = process_image(payload)
    elif mode == "video":
        result = process_video(payload)
    elif mode == "report_latest":
        result = report_latest(payload)
    elif mode == "report_daily_pdf":
        result = report_daily(payload, "pdf")
    elif mode == "report_daily_xlsx":
        result = report_daily(payload, "xlsx")
    else:
        result = {"ok": False, "error": "Invalid mode"}

    print(json.dumps(result))


if __name__ == "__main__":
    main()
