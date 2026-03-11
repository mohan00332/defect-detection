import os
import threading
import json
import csv
from datetime import datetime, timedelta

import cv2
import numpy as np
import requests
from flask import Flask, render_template, request, jsonify, Response, send_from_directory, send_file, redirect, url_for
from ultralytics import YOLO
import Nut_Detection
import Bolt_Detection
import Gear_Detection

try:
    import gspread
    from oauth2client.service_account import ServiceAccountCredentials
except Exception:
    gspread = None
    ServiceAccountCredentials = None

try:
    from fpdf import FPDF
except Exception:
    FPDF = None

try:
    from openpyxl import Workbook
except Exception:
    Workbook = None

APP_ROOT = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(APP_ROOT, "uploads")
OUTPUT_DIR = os.path.join(APP_ROOT, "outputs")
REPORT_DIR = os.path.join(APP_ROOT, "reports")
GOOGLE_DIR = os.path.join(APP_ROOT, "google")
MODEL_DIR = os.path.join(APP_ROOT, "models")
IMAGES_DIR = os.path.join(APP_ROOT, "images")

for d in (UPLOAD_DIR, OUTPUT_DIR, REPORT_DIR, GOOGLE_DIR, MODEL_DIR, IMAGES_DIR):
    os.makedirs(d, exist_ok=True)

app = Flask(__name__)
app.config["UPLOAD_FOLDER"] = UPLOAD_DIR

CATEGORY_TO_MODEL = {
    "All": None,
    "Nut": "Nut.pt",
    "Bolt": "Bolt.pt",
    "Gear": "Gear.pt",
}

CATEGORY_CONF = {
    "All": 0.25,
    "Nut": 0.5,
    "Bolt": 0.2,
    "Gear": 0.2,
}

VALID_CATEGORIES = {"All", "Nut", "Bolt", "Gear"}

CAMERA_INDEX = int(os.environ.get("CAMERA_INDEX", "0"))
SHEET_NAME = os.environ.get("SHEET_NAME", "Defect_Detections")

_model_cache = {}
_camera_lock = threading.Lock()
_camera_cap = None
_camera_active = False

_last_live_result = None
_last_report = None
_live_session = {
    "category": None,
    "max_detected": 0,
    "max_defect": 0,
    "last_detected": 0,
    "last_defect": 0,
    "last_good": 0,
    "frame": None,
    "max_per": None,
    "last_per": None,
}

_stats = {
    "Nut": {"detected": 0, "defect": 0, "good": 0},
    "Bolt": {"detected": 0, "defect": 0, "good": 0},
    "Gear": {"detected": 0, "defect": 0, "good": 0},
}
_stats_meta = {"last_reset": None}

STATS_PATH = os.path.join(APP_ROOT, "stats.json")
LOG_PATH = os.path.join(REPORT_DIR, "detections_log.csv")
SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_SERVICE_ROLE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
SUPABASE_ANON_KEY = os.environ.get("SUPABASE_ANON_KEY")
SUPABASE_TABLE = os.environ.get("SUPABASE_TABLE", "detections")


def _log_to_supabase(payload):
    key = SUPABASE_SERVICE_ROLE_KEY or SUPABASE_ANON_KEY
    if not SUPABASE_URL or not key:
        return
    try:
        url = f"{SUPABASE_URL}/rest/v1/{SUPABASE_TABLE}"
        headers = {
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal",
        }
        requests.post(url, headers=headers, json=payload, timeout=5)
    except Exception:
        pass


def _load_stats():
    if not os.path.exists(STATS_PATH):
        return
    try:
        with open(STATS_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        meta = data.get("_meta") if isinstance(data, dict) else None
        if isinstance(meta, dict):
            _stats_meta["last_reset"] = meta.get("last_reset")
        for key in _stats:
            if key in data:
                _stats[key]["detected"] = int(data[key].get("detected", 0))
                _stats[key]["defect"] = int(data[key].get("defect", 0))
                _stats[key]["good"] = int(data[key].get("good", 0))
    except Exception:
        pass


def _save_stats():
    try:
        with open(STATS_PATH, "w", encoding="utf-8") as f:
            json.dump({"_meta": _stats_meta, **_stats}, f)
    except Exception:
        pass


def _maybe_reset_stats():
    today = datetime.now().date().isoformat()
    if _stats_meta.get("last_reset") == today:
        return False
    for key in _stats:
        _stats[key]["detected"] = 0
        _stats[key]["defect"] = 0
        _stats[key]["good"] = 0
    _stats_meta["last_reset"] = today
    _save_stats()
    return True


_load_stats()
if _stats_meta.get("last_reset") is None:
    _stats_meta["last_reset"] = datetime.now().date().isoformat()
    _save_stats()
else:
    _maybe_reset_stats()


def _resolve_model_path(model_name: str) -> str:
    model_path = os.path.join(MODEL_DIR, model_name)
    if os.path.exists(model_path):
        return model_path
    fallback = os.path.join(APP_ROOT, model_name)
    return fallback


def get_model(category: str) -> YOLO:
    if category not in CATEGORY_TO_MODEL:
        raise ValueError("Unsupported category")
    if category == "All":
        raise ValueError("No single model for All category")
    if category in _model_cache:
        return _model_cache[category]
    model_path = _resolve_model_path(CATEGORY_TO_MODEL[category])
    model = YOLO(model_path)
    _model_cache[category] = model
    return model


def is_defect(class_name: str) -> bool:
    name = class_name.lower().strip()
    if "non_defect" in name or "non-defect" in name or "nondefect" in name:
        return False
    if "gear_defect" in name:
        return True
    return ("defect" in name) or ("bad" in name) or ("fault" in name)


def preprocess_frame(category: str, frame):
    if category == "Gear":
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        blur = cv2.GaussianBlur(gray, (5, 5), 0)
        edges = cv2.Canny(blur, 50, 150)
        edges_3ch = cv2.cvtColor(edges, cv2.COLOR_GRAY2BGR)
        return edges_3ch
    return frame


def run_detection_on_frame(category: str, frame, conf_override=None, iou_override=None, label_mode="confidence"):
    if category == "Nut":
        frame, detected, defect = Nut_Detection.detect_frame(frame, conf_override, iou_override, label_mode)
        return frame, detected, defect
    if category == "Bolt":
        frame, detected, defect = Bolt_Detection.detect_frame(frame, conf_override, iou_override, label_mode)
        return frame, detected, defect
    if category == "Gear":
        frame, detected, defect, _ = Gear_Detection.detect_frame(frame, conf_override, iou_override, label_mode)
        return frame, detected, defect

    if category == "All":
        raise ValueError("Use run_detection_all for All category")

    model = get_model(category)
    conf = CATEGORY_CONF.get(category, 0.25)
    model_input = preprocess_frame(category, frame)
    results = model(model_input, conf=conf, iou=iou_override) if iou_override is not None else model(model_input, conf=conf)

    detected = 0
    defect = 0

    for box in results[0].boxes:
        detected += 1
        x1, y1, x2, y2 = map(int, box.xyxy[0])
        conf_score = float(box.conf[0])
        cls_id = int(box.cls[0])
        class_name = model.names[cls_id]

        label = f"{class_name} {conf_score:.2f}"
        if is_defect(class_name):
            defect += 1
            color = (0, 0, 255)
        else:
            color = (0, 255, 0)

        cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
        cv2.putText(frame, label, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)

    good = max(detected - defect, 0)
    cv2.putText(frame, f"Defect: {defect}", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2)
    cv2.putText(frame, f"Non-Defect: {good}", (10, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)

    return frame, detected, defect


def run_detection_all(frame, conf_override=None, iou_override=None, label_mode="confidence"):
    out = frame
    n_frame, n_det, n_def = Nut_Detection.detect_frame(out, conf_override, iou_override, label_mode, draw_counts=False)
    out = n_frame
    b_frame, b_det, b_def = Bolt_Detection.detect_frame(out, conf_override, iou_override, label_mode, draw_counts=False)
    out = b_frame
    g_frame, g_det, g_def, _ = Gear_Detection.detect_frame(out, conf_override, iou_override, label_mode, draw_counts=False)
    out = g_frame

    totals = (n_det + b_det + g_det, n_def + b_def + g_def)
    good = max(totals[0] - totals[1], 0)
    cv2.putText(out, f"Defect: {totals[1]}", (10, 30),
                cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2)
    cv2.putText(out, f"Non-Defect: {good}", (10, 60),
                cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)
    per_category = {
        "Nut": (n_det, n_def),
        "Bolt": (b_det, b_def),
        "Gear": (g_det, g_def),
    }
    return out, totals, per_category


def _timestamp():
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


def _append_to_sheet(row):
    if gspread is None or ServiceAccountCredentials is None:
        return False, "gspread not available"

    creds_path = os.path.join(GOOGLE_DIR, "credentials.json")
    if not os.path.exists(creds_path):
        return False, "google/credentials.json not found"

    scope = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
    ]

    try:
        creds = ServiceAccountCredentials.from_json_keyfile_name(creds_path, scope)
        client = gspread.authorize(creds)
        sheet = client.open(SHEET_NAME).sheet1
        sheet.append_row(row)
        return True, "ok"
    except Exception as e:
        return False, str(e)


def _update_stats(category: str, detected: int, defect: int):
    _maybe_reset_stats()
    good = max(detected - defect, 0)
    _stats[category]["detected"] += detected
    _stats[category]["defect"] += defect
    _stats[category]["good"] += good
    _save_stats()
    return good


def _update_stats_many(per_category):
    for cat, (detected, defect) in per_category.items():
        _update_stats(cat, detected, defect)


def _build_report(data, image_path=None):
    if FPDF is None:
        return None

    filename = f"report_{_timestamp()}.pdf"
    report_path = os.path.join(REPORT_DIR, filename)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Manufacturing Defect Inspection Report", ln=True)

    pdf.set_font("Helvetica", size=12)
    pdf.ln(4)

    lines = [
        f"Date: {data['date']}",
        f"Time: {data['time']}",
        f"Category: {data['category']}",
        f"Expected Count: {data['expected_count']}",
        f"Detected Count: {data['detected_count']}",
        f"Defect Count: {data['defect_count']}",
        f"Good Count: {data['good_count']}",
    ]

    for line in lines:
        pdf.cell(0, 8, line, ln=True)

    if image_path and os.path.exists(image_path):
        pdf.ln(6)
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, "Detection Preview", ln=True)
        pdf.ln(2)
        pdf.image(image_path, w=170)

    pdf.output(report_path)
    return report_path


def _record_detection(category, expected_count, detected, defect, good, output_image_path=None):
    now = datetime.now()
    row = [
        now.strftime("%Y-%m-%d"),
        now.strftime("%H:%M:%S"),
        category,
        expected_count,
        detected,
        defect,
        good,
    ]

    _append_to_sheet(row)
    _append_to_local_log(row)
    _log_to_supabase({
        "timestamp": now.isoformat(),
        "category": category,
        "expected_count": expected_count,
        "detected": detected,
        "defect": defect,
        "good": good,
    })

    data = {
        "date": now.strftime("%Y-%m-%d"),
        "time": now.strftime("%H:%M:%S"),
        "category": category,
        "expected_count": expected_count,
        "detected_count": detected,
        "defect_count": defect,
        "good_count": good,
    }

    report_path = _build_report(data, output_image_path)
    return data, report_path


def _append_to_local_log(row):
    header = ["date", "time", "category", "expected_count", "detected", "defect", "good"]
    file_exists = os.path.exists(LOG_PATH)
    try:
        with open(LOG_PATH, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if not file_exists:
                writer.writerow(header)
            writer.writerow(row)
    except Exception:
        pass


def _read_log_rows():
    if not os.path.exists(LOG_PATH):
        return []
    rows = []
    try:
        with open(LOG_PATH, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(row)
    except Exception:
        return []
    return rows


def _save_image(frame, prefix):
    filename = f"{prefix}_{_timestamp()}.jpg"
    path = os.path.join(OUTPUT_DIR, filename)
    cv2.imwrite(path, frame)
    return path, filename


def _decode_image_bytes(file_bytes):
    np_arr = np.frombuffer(file_bytes, np.uint8)
    return cv2.imdecode(np_arr, cv2.IMREAD_COLOR)


def _safe_int(value, default=0):
    try:
        return int(value)
    except Exception:
        return default


def _safe_float(value, default=None):
    try:
        return float(value)
    except Exception:
        return default


def _normalize_label_mode(value):
    if value in {"confidence", "label", "none"}:
        return value
    return "confidence"


def _is_image_ext(ext: str) -> bool:
    return ext in {".jpg", ".jpeg", ".png"}


def _is_video_ext(ext: str) -> bool:
    return ext in {".mp4", ".avi"}


def _is_image_mime(mime: str) -> bool:
    return (mime or "").startswith("image/")


def _is_video_mime(mime: str) -> bool:
    return (mime or "").startswith("video/")


def generate_frames(category: str):
    global _camera_cap, _camera_active, _last_live_result, _last_report, _live_session

    with _camera_lock:
        if _camera_cap is None:
            _camera_cap = cv2.VideoCapture(CAMERA_INDEX, cv2.CAP_DSHOW)

    if _live_session["category"] != category:
        _live_session = {
            "category": category,
            "max_detected": 0,
            "max_defect": 0,
            "last_detected": 0,
            "last_defect": 0,
            "last_good": 0,
            "frame": None,
            "max_per": None,
            "last_per": None,
            "conf": _live_session.get("conf"),
            "iou": _live_session.get("iou"),
            "label_mode": _live_session.get("label_mode", "confidence"),
        }

    while True:
        with _camera_lock:
            if not _camera_active or _camera_cap is None:
                break
            cap = _camera_cap

        ret, frame = cap.read()
        if not ret:
            break

        if category == "All":
            processed, (detected, defect), per_category = run_detection_all(
                frame, _live_session.get("conf"), _live_session.get("iou"), _live_session.get("label_mode", "confidence")
            )
            good = max(detected - defect, 0)
        else:
            processed, detected, defect = run_detection_on_frame(
                category, frame, _live_session.get("conf"), _live_session.get("iou"), _live_session.get("label_mode", "confidence")
            )
            good = max(detected - defect, 0)

        _last_live_result = {
            "category": category,
            "detected": detected,
            "defect": defect,
            "good": good,
            "frame": processed.copy(),
        }
        _live_session["last_detected"] = detected
        _live_session["last_defect"] = defect
        _live_session["last_good"] = good
        _live_session["frame"] = processed.copy()
        if detected > _live_session["max_detected"]:
            _live_session["max_detected"] = detected
        if defect > _live_session["max_defect"]:
            _live_session["max_defect"] = defect
        if category == "All":
            _live_session["last_per"] = per_category
            if _live_session["max_per"] is None:
                _live_session["max_per"] = {
                    "Nut": {"detected": 0, "defect": 0},
                    "Bolt": {"detected": 0, "defect": 0},
                    "Gear": {"detected": 0, "defect": 0},
                }
            for cat, (det, dft) in per_category.items():
                if det > _live_session["max_per"][cat]["detected"]:
                    _live_session["max_per"][cat]["detected"] = det
                if dft > _live_session["max_per"][cat]["defect"]:
                    _live_session["max_per"][cat]["defect"] = dft

        ret2, buffer = cv2.imencode(".jpg", processed)
        if not ret2:
            continue

        frame_bytes = buffer.tobytes()
        yield (b"--frame\r\n"
               b"Content-Type: image/jpeg\r\n\r\n" + frame_bytes + b"\r\n")


def _get_sample_images():
    images = []
    try:
        for name in sorted(os.listdir(IMAGES_DIR)):
            ext = os.path.splitext(name)[1].lower()
            if ext in {".jpg", ".jpeg", ".png"}:
                images.append({
                    "name": name,
                    "url": f"/images/{name}"
                })
    except Exception:
        pass
    return images


@app.context_processor
def inject_sample_images():
    return {"sample_images": _get_sample_images()}


@app.route("/health")
def health():
    return jsonify({"ok": True})


@app.route("/")
def index():
    return render_template("index.html", sample_images=_get_sample_images())


@app.route("/samples")
def samples():
    return redirect(url_for("samples_view"))


@app.route("/samples-view")
def samples_view():
    return render_template("samples.html", sample_images=_get_sample_images())


@app.route("/samples-json")
def samples_json():
    return jsonify({"ok": True, "images": _get_sample_images()})


@app.route("/start_camera", methods=["POST"])
def start_camera():
    global _camera_active, _live_session
    data = request.get_json(force=True)
    category = data.get("category")
    conf_override = _safe_float(data.get("conf"), None)
    iou_override = _safe_float(data.get("iou"), None)
    label_mode = _normalize_label_mode(data.get("label_mode"))

    if category not in VALID_CATEGORIES:
        return jsonify({"ok": False, "error": "Invalid category"}), 400

    _live_session = {
        "category": category,
        "max_detected": 0,
        "max_defect": 0,
        "last_detected": 0,
        "last_defect": 0,
        "last_good": 0,
        "frame": None,
        "max_per": None,
        "last_per": None,
        "conf": conf_override,
        "iou": iou_override,
        "label_mode": label_mode,
    }
    _camera_active = True
    return jsonify({"ok": True})


@app.route("/stop_camera", methods=["POST"])
def stop_camera():
    global _camera_active, _camera_cap, _last_live_result, _last_report, _live_session

    data = request.get_json(force=True)
    expected = _safe_int(data.get("expected_count"), 0)

    _camera_active = False

    with _camera_lock:
        if _camera_cap is not None:
            _camera_cap.release()
            _camera_cap = None

    if _last_live_result:
        category = _last_live_result["category"]
        frame = _live_session["frame"] or _last_live_result["frame"]

        if category == "All" and _live_session["max_per"]:
            detected = sum(v["detected"] for v in _live_session["max_per"].values())
            defect = sum(v["defect"] for v in _live_session["max_per"].values())
            good = max(detected - defect, 0)
            output_path, output_name = _save_image(frame, "live_all")
            per_category = {
                cat: (vals["detected"], vals["defect"])
                for cat, vals in _live_session["max_per"].items()
            }
            _update_stats_many(per_category)
            data_row, report_path = _record_detection(category, expected, detected, defect, good, output_path)
        else:
            detected = _live_session["max_detected"]
            defect = _live_session["max_defect"]
            good = max(detected - defect, 0)
            output_path, output_name = _save_image(frame, f"live_{category.lower()}")
            _update_stats(category, detected, defect)
            data_row, report_path = _record_detection(category, expected, detected, defect, good, output_path)

        _last_report = report_path

        return jsonify({
            "ok": True,
            "category": category,
            "detected": detected,
            "defect": defect,
            "good": good,
            "output_url": f"/outputs/{output_name}",
        })

    return jsonify({"ok": True, "message": "Camera stopped"})


@app.route("/live_stats")
def live_stats():
    if _last_live_result is None:
        return jsonify({"ok": False, "error": "No live data"}), 404
    return jsonify({
        "ok": True,
        "category": _last_live_result["category"],
        "detected": _live_session["last_detected"],
        "defect": _live_session["last_defect"],
        "good": _live_session["last_good"],
        "non_defect": _live_session["last_good"],
        "max_detected": _live_session["max_detected"],
        "max_defect": _live_session["max_defect"],
    })


@app.route("/video_feed")
def video_feed():
    category = request.args.get("category", "Nut")
    if category not in VALID_CATEGORIES:
        category = "Nut"
    return Response(generate_frames(category), mimetype="multipart/x-mixed-replace; boundary=frame")


@app.route("/detect_image", methods=["POST"])
def detect_image():
    global _last_report

    if "image" not in request.files:
        return jsonify({"ok": False, "error": "No image uploaded"}), 400

    category = request.form.get("category", "Nut")
    expected = _safe_int(request.form.get("expected_count"), 0)
    conf_override = _safe_float(request.form.get("conf"), None)
    iou_override = _safe_float(request.form.get("iou"), None)
    label_mode = _normalize_label_mode(request.form.get("label_mode"))

    if category not in VALID_CATEGORIES:
        return jsonify({"ok": False, "error": "Invalid category"}), 400

    file = request.files["image"]
    file_bytes = file.read()
    frame = _decode_image_bytes(file_bytes)
    if frame is None:
        return jsonify({"ok": False, "error": "Invalid image"}), 400

    if category == "All":
        processed, (detected, defect), per_category = run_detection_all(
            frame, conf_override, iou_override, label_mode
        )
        good = max(detected - defect, 0)
        output_path, output_name = _save_image(processed, "image_all")
        _update_stats_many(per_category)
    else:
        processed, detected, defect = run_detection_on_frame(
            category, frame, conf_override, iou_override, label_mode
        )
        good = max(detected - defect, 0)
        output_path, output_name = _save_image(processed, f"image_{category.lower()}")
        _update_stats(category, detected, defect)

    data_row, report_path = _record_detection(category, expected, detected, defect, good, output_path)
    _last_report = report_path

    return jsonify({
        "ok": True,
        "category": category,
        "detected": detected,
        "defect": defect,
        "good": good,
        "output_url": f"/outputs/{output_name}",
    })


@app.route("/detect_image_url", methods=["POST"])
def detect_image_url():
    global _last_report

    data = request.get_json(force=True)
    url = data.get("url", "").strip()
    category = data.get("category", "Nut")
    expected = _safe_int(data.get("expected_count"), 0)
    conf_override = _safe_float(data.get("conf"), None)
    iou_override = _safe_float(data.get("iou"), None)
    label_mode = _normalize_label_mode(data.get("label_mode"))

    if not url.startswith("http://") and not url.startswith("https://"):
        return jsonify({"ok": False, "error": "Invalid URL"}), 400

    if category not in VALID_CATEGORIES:
        return jsonify({"ok": False, "error": "Invalid category"}), 400

    try:
        resp = requests.get(url, timeout=10)
        if resp.status_code != 200:
            return jsonify({"ok": False, "error": "Failed to fetch image"}), 400
        content = resp.content
        if len(content) > 5 * 1024 * 1024:
            return jsonify({"ok": False, "error": "Image too large"}), 400
    except Exception:
        return jsonify({"ok": False, "error": "Failed to fetch image"}), 400

    frame = _decode_image_bytes(content)
    if frame is None:
        return jsonify({"ok": False, "error": "Invalid image"}), 400

    if category == "All":
        processed, (detected, defect), per_category = run_detection_all(
            frame, conf_override, iou_override, label_mode
        )
        good = max(detected - defect, 0)
        output_path, output_name = _save_image(processed, "url_all")
        _update_stats_many(per_category)
    else:
        processed, detected, defect = run_detection_on_frame(
            category, frame, conf_override, iou_override, label_mode
        )
        good = max(detected - defect, 0)
        output_path, output_name = _save_image(processed, f"url_{category.lower()}")
        _update_stats(category, detected, defect)

    data_row, report_path = _record_detection(category, expected, detected, defect, good, output_path)
    _last_report = report_path

    return jsonify({
        "ok": True,
        "category": category,
        "detected": detected,
        "defect": defect,
        "good": good,
        "output_url": f"/outputs/{output_name}",
    })


@app.route("/detect_video", methods=["POST"])
def detect_video():
    global _last_report

    if "video" not in request.files:
        return jsonify({"ok": False, "error": "No video uploaded"}), 400

    category = request.form.get("category", "Nut")
    expected = _safe_int(request.form.get("expected_count"), 0)

    if category not in VALID_CATEGORIES:
        return jsonify({"ok": False, "error": "Invalid category"}), 400

    file = request.files["video"]
    filename = f"video_{category.lower()}_{_timestamp()}.mp4"
    upload_path = os.path.join(UPLOAD_DIR, filename)
    file.save(upload_path)

    cap = cv2.VideoCapture(upload_path)
    if not cap.isOpened():
        return jsonify({"ok": False, "error": "Failed to open video"}), 400

    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    fps = cap.get(cv2.CAP_PROP_FPS) or 24

    output_name = f"output_{category.lower()}_{_timestamp()}.mp4"
    output_path = os.path.join(OUTPUT_DIR, output_name)

    fourcc = cv2.VideoWriter_fourcc(*"mp4v")
    writer = cv2.VideoWriter(output_path, fourcc, fps, (width, height))

    detected_total = 0
    defect_total = 0
    per_category_totals = {
        "Nut": {"detected": 0, "defect": 0},
        "Bolt": {"detected": 0, "defect": 0},
        "Gear": {"detected": 0, "defect": 0},
    }
    last_frame = None

    while True:
        ret, frame = cap.read()
        if not ret:
            break
        if category == "All":
            processed, (detected, defect), per_category = run_detection_all(frame)
            for cat, (det, dft) in per_category.items():
                per_category_totals[cat]["detected"] += det
                per_category_totals[cat]["defect"] += dft
        else:
            processed, detected, defect = run_detection_on_frame(category, frame)
        detected_total += detected
        defect_total += defect
        writer.write(processed)
        last_frame = processed

    cap.release()
    writer.release()

    good = max(detected_total - defect_total, 0)
    if category == "All":
        per_category = {cat: (v["detected"], v["defect"]) for cat, v in per_category_totals.items()}
        _update_stats_many(per_category)
    else:
        _update_stats(category, detected_total, defect_total)

    output_image_path = None
    if last_frame is not None:
        output_image_path, _ = _save_image(last_frame, f"video_{category.lower()}")

    data_row, report_path = _record_detection(category, expected, detected_total, defect_total, good, output_image_path)
    _last_report = report_path

    return jsonify({
        "ok": True,
        "category": category,
        "detected": detected_total,
        "defect": defect_total,
        "good": good,
        "output_url": f"/outputs/{output_name}",
    })


@app.route("/upload", methods=["POST"])
def upload():
    global _last_report

    if "file" not in request.files:
        return jsonify({"ok": False, "error": "No file uploaded"}), 400

    category = request.form.get("category", "Nut")
    expected = _safe_int(request.form.get("expected_count"), 0)
    conf_override = _safe_float(request.form.get("conf"), None)
    iou_override = _safe_float(request.form.get("iou"), None)
    label_mode = _normalize_label_mode(request.form.get("label_mode"))
    file = request.files["file"]

    if category not in VALID_CATEGORIES:
        return jsonify({"ok": False, "error": "Invalid category"}), 400

    filename = file.filename or ""
    ext = os.path.splitext(filename)[1].lower()
    mime = file.mimetype or ""

    if _is_image_ext(ext) or _is_image_mime(mime):
        file_bytes = file.read()
        frame = _decode_image_bytes(file_bytes)
        if frame is None:
            return jsonify({"ok": False, "error": "Invalid image"}), 400

        if category == "All":
            processed, (detected, defect), per_category = run_detection_all(
                frame, conf_override, iou_override, label_mode
            )
            good = max(detected - defect, 0)
            output_path, output_name = _save_image(processed, "image_all")
            _update_stats_many(per_category)
        else:
            processed, detected, defect = run_detection_on_frame(
                category, frame, conf_override, iou_override, label_mode
            )
            good = max(detected - defect, 0)
            output_path, output_name = _save_image(processed, f"image_{category.lower()}")
            _update_stats(category, detected, defect)

        data_row, report_path = _record_detection(category, expected, detected, defect, good, output_path)
        _last_report = report_path

        return jsonify({
            "ok": True,
            "category": category,
            "detected": detected,
            "defect": defect,
            "good": good,
            "output_url": f"/outputs/{output_name}",
        })

    if _is_video_ext(ext) or _is_video_mime(mime):
        filename = f"video_{category.lower()}_{_timestamp()}{ext or '.mp4'}"
        upload_path = os.path.join(UPLOAD_DIR, filename)
        file.save(upload_path)

        cap = cv2.VideoCapture(upload_path)
        if not cap.isOpened():
            return jsonify({"ok": False, "error": "Failed to open video"}), 400

        width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        fps = cap.get(cv2.CAP_PROP_FPS) or 24

        output_name = f"output_{category.lower()}_{_timestamp()}.mp4"
        output_path = os.path.join(OUTPUT_DIR, output_name)

        fourcc = cv2.VideoWriter_fourcc(*"mp4v")
        writer = cv2.VideoWriter(output_path, fourcc, fps, (width, height))

        detected_total = 0
        defect_total = 0
        per_category_totals = {
            "Nut": {"detected": 0, "defect": 0},
            "Bolt": {"detected": 0, "defect": 0},
            "Gear": {"detected": 0, "defect": 0},
        }
        last_frame = None

        while True:
            ret, frame = cap.read()
            if not ret:
                break
            if category == "All":
                processed, (detected, defect), per_category = run_detection_all(frame)
                for cat, (det, dft) in per_category.items():
                    per_category_totals[cat]["detected"] += det
                    per_category_totals[cat]["defect"] += dft
            else:
                processed, detected, defect = run_detection_on_frame(category, frame)
            detected_total += detected
            defect_total += defect
            writer.write(processed)
            last_frame = processed

        cap.release()
        writer.release()

        good = max(detected_total - defect_total, 0)
        if category == "All":
            per_category = {cat: (v["detected"], v["defect"]) for cat, v in per_category_totals.items()}
            _update_stats_many(per_category)
        else:
            _update_stats(category, detected_total, defect_total)

        output_image_path = None
        if last_frame is not None:
            output_image_path, _ = _save_image(last_frame, f"video_{category.lower()}")

        data_row, report_path = _record_detection(category, expected, detected_total, defect_total, good, output_image_path)
        _last_report = report_path

        return jsonify({
            "ok": True,
            "category": category,
            "detected": detected_total,
            "defect": defect_total,
            "good": good,
            "output_url": f"/outputs/{output_name}",
        })

    return jsonify({"ok": False, "error": "Unsupported file type"}), 400


@app.route("/stats")
def stats():
    _maybe_reset_stats()
    return jsonify({"ok": True, "stats": _stats})


@app.route("/reset_stats", methods=["POST"])
def reset_stats():
    for key in _stats:
        _stats[key]["detected"] = 0
        _stats[key]["defect"] = 0
        _stats[key]["good"] = 0
    _stats_meta["last_reset"] = datetime.now().date().isoformat()
    _save_stats()
    return jsonify({"ok": True, "stats": _stats})


@app.route("/outputs/<path:filename>")
def outputs(filename):
    return send_from_directory(OUTPUT_DIR, filename)


@app.route("/uploads/<path:filename>")
def uploads(filename):
    return send_from_directory(UPLOAD_DIR, filename)


@app.route("/reports/<path:filename>")
def reports(filename):
    return send_from_directory(REPORT_DIR, filename)


@app.route("/images/<path:filename>")
def images(filename):
    return send_from_directory(IMAGES_DIR, filename)


@app.route("/report/latest")
def report_latest():
    if _last_report and os.path.exists(_last_report):
        return send_file(_last_report, as_attachment=True)
    return jsonify({"ok": False, "error": "No report generated yet"}), 404


@app.route("/report/daily.csv")
def report_daily_csv():
    date_str = request.args.get("date")
    if not date_str:
        date_str = datetime.now().strftime("%Y-%m-%d")

    output_name = f"daily_report_{date_str}.csv"
    output_path = os.path.join(REPORT_DIR, output_name)

    header = ["date", "time", "category", "expected_count", "detected", "defect", "good"]
    rows = []
    for row in _read_log_rows():
        if row.get("date") == date_str:
            rows.append([
                row.get("date", ""),
                row.get("time", ""),
                row.get("category", ""),
                row.get("expected_count", ""),
                row.get("detected", ""),
                row.get("defect", ""),
                row.get("good", ""),
            ])

    try:
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(header)
            writer.writerows(rows)
    except Exception:
        return jsonify({"ok": False, "error": "Failed to create report"}), 500

    return send_file(output_path, as_attachment=True)


@app.route("/report/daily.pdf")
def report_daily_pdf():
    if FPDF is None:
        return jsonify({"ok": False, "error": "FPDF not available"}), 500

    date_str = request.args.get("date")
    if not date_str:
        date_str = datetime.now().strftime("%Y-%m-%d")

    rows = [row for row in _read_log_rows() if row.get("date") == date_str]

    output_name = f"daily_report_{date_str}.pdf"
    output_path = os.path.join(REPORT_DIR, output_name)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, f"Daily Defect Report - {date_str}", ln=True)
    pdf.ln(4)

    pdf.set_font("Helvetica", "B", 10)
    col_widths = [24, 20, 18, 32, 22, 18, 18]
    headers = ["Date", "Time", "Cat", "Expected", "Detected", "Defect", "Good"]
    for w, h in zip(col_widths, headers):
        pdf.cell(w, 8, h, border=1)
    pdf.ln()

    pdf.set_font("Helvetica", size=10)
    if rows:
        for row in rows:
            values = [
                row.get("date", ""),
                row.get("time", ""),
                row.get("category", ""),
                str(row.get("expected_count", "")),
                str(row.get("detected", "")),
                str(row.get("defect", "")),
                str(row.get("good", "")),
            ]
            for w, v in zip(col_widths, values):
                pdf.cell(w, 8, v, border=1)
            pdf.ln()
    else:
        pdf.cell(0, 8, "No detections found for this date.", ln=True)

    pdf.output(output_path)
    return send_file(output_path, as_attachment=True)


@app.route("/report/daily.xlsx")
def report_daily_xlsx():
    if Workbook is None:
        return jsonify({"ok": False, "error": "openpyxl not available"}), 500

    date_str = request.args.get("date")
    if not date_str:
        date_str = datetime.now().strftime("%Y-%m-%d")

    rows = [row for row in _read_log_rows() if row.get("date") == date_str]

    output_name = f"daily_report_{date_str}.xlsx"
    output_path = os.path.join(REPORT_DIR, output_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Report"

    header = ["date", "time", "category", "expected_count", "detected", "defect", "good"]
    ws.append(header)
    for row in rows:
        ws.append([
            row.get("date", ""),
            row.get("time", ""),
            row.get("category", ""),
            row.get("expected_count", ""),
            row.get("detected", ""),
            row.get("defect", ""),
            row.get("good", ""),
        ])

    wb.save(output_path)
    return send_file(output_path, as_attachment=True)


@app.route("/model_info")
def model_info():
    info = {}
    for category in ("Nut", "Bolt", "Gear"):
        try:
            model = get_model(category)
            names = model.names or {}
            info[category] = [str(names[k]) for k in sorted(names.keys())]
        except Exception as e:
            info[category] = {"error": str(e)}
    return jsonify({"ok": True, "models": info})


@app.route("/analytics/summary")
def analytics_summary():
    today = datetime.now().date()
    days = [today - timedelta(days=i) for i in range(6, -1, -1)]
    day_keys = [d.strftime("%Y-%m-%d") for d in days]
    trend = {k: {"defect": 0, "good": 0, "detected": 0} for k in day_keys}

    for row in _read_log_rows():
        d = row.get("date")
        if d not in trend:
            continue
        defect = _safe_int(row.get("defect"), 0)
        good = _safe_int(row.get("good"), 0)
        detected = _safe_int(row.get("detected"), 0)
        trend[d]["defect"] += defect
        trend[d]["good"] += good
        trend[d]["detected"] += detected

    trend_list = [{"date": k, **trend[k]} for k in day_keys]

    categories = {
        "Nut": _stats.get("Nut", {"detected": 0, "defect": 0, "good": 0}),
        "Bolt": _stats.get("Bolt", {"detected": 0, "defect": 0, "good": 0}),
        "Gear": _stats.get("Gear", {"detected": 0, "defect": 0, "good": 0}),
    }
    totals = {
        "detected": sum(v["detected"] for v in categories.values()),
        "defect": sum(v["defect"] for v in categories.values()),
        "good": sum(v["good"] for v in categories.values()),
    }

    return jsonify({
        "ok": True,
        "trend": trend_list,
        "categories": categories,
        "totals": totals,
    })


def _add_api_alias(rule, view_func, methods=None):
    endpoint = f"api_{view_func.__name__}_{rule.strip('/').replace('/', '_') or 'root'}"
    app.add_url_rule(f"/api{rule}", endpoint=endpoint, view_func=view_func, methods=methods)


_add_api_alias("/health", health, ["GET"])
_add_api_alias("/samples-json", samples_json, ["GET"])
_add_api_alias("/start_camera", start_camera, ["POST"])
_add_api_alias("/stop_camera", stop_camera, ["POST"])
_add_api_alias("/live_stats", live_stats, ["GET"])
_add_api_alias("/video_feed", video_feed, ["GET"])
_add_api_alias("/detect_image", detect_image, ["POST"])
_add_api_alias("/detect_image_url", detect_image_url, ["POST"])
_add_api_alias("/detect_video", detect_video, ["POST"])
_add_api_alias("/upload", upload, ["POST"])
_add_api_alias("/stats", stats, ["GET"])
_add_api_alias("/reset_stats", reset_stats, ["POST"])
_add_api_alias("/report/latest", report_latest, ["GET"])
_add_api_alias("/report/daily.csv", report_daily_csv, ["GET"])
_add_api_alias("/report/daily.pdf", report_daily_pdf, ["GET"])
_add_api_alias("/report/daily.xlsx", report_daily_xlsx, ["GET"])
_add_api_alias("/analytics/summary", analytics_summary, ["GET"])
_add_api_alias("/model_info", model_info, ["GET"])


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
